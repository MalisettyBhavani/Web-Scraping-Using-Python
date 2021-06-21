from bs4 import BeautifulSoup
import urllib.request,urllib.error,urllib.parse
import ssl
import openpyxl

ctx=ssl.create_default_context()
ctx.check_hostname= False
ctx.verify_mode=ssl.CERT_NONE


page=urllib.request.urlopen("https://www.flipkart.com/search?q=books&otracker=search&otracker1=search&marketplace=FLIPKART&as-show=on&as=off&as-pos=1&as-type=HISTORY")
soup=BeautifulSoup(page,'html.parser')

booknames=soup.select("a.s1Q9rs")

price_current=soup.select("a._8VNy32 div._25b18c ._30jeq3")

price_before=soup.select("a._8VNy32 div._25b18c ._3I9_wc")

ratings=soup.select("div.gUuXy-._2D5lwg div._3LWZlK")

reviews=soup.select("div.gUuXy-._2D5lwg span._2_R_DZ")

##Storing the details of the books in the excel sheet ...

wb=openpyxl.load_workbook("F:\projects\Flipkart_Books.xlsx")
sheet1=wb["Books"]

rows=1
for i in booknames:
    sheet1.cell(rows+1,1).value=str(i.text)
    rows+=1

rows=1
for i in price_current:
    sheet1.cell(rows+1,2).value=str(i.text)
    rows+=1

rows=1
for i in price_before:
    sheet1.cell(rows+1,3).value=str(i.text)
    rows+=1

rows=1
for i in ratings:
    sheet1.cell(rows+1,4).value=str(i.text)
    rows+=1

rows=1
for i in reviews:
    sheet1.cell(rows+1,5).value=str(i.text)[1:-1]
    rows+=1

wb.save("F:\projects\Flipkart_Books.xlsx")
wb.close()